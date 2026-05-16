import streamlit as st
import pdfplumber
import re
import os
from openpyxl import load_workbook
from io import BytesIO

# ---------------------------------------------------
# PAGE CONFIG
# ---------------------------------------------------

st.set_page_config(
    page_title="DISCOM Bill Analysis",
    layout="wide"
)

st.title("⚡ DISCOM Bill Analysis App")
st.subheader("Before Solar vs After Solar Analysis")

# ---------------------------------------------------
# FIXED VALUES
# ---------------------------------------------------

solar_capacity = 1000
plant_load = 1800

# ---------------------------------------------------
# MANUAL INPUTS
# ---------------------------------------------------

st.markdown("## ⚡ Manual Inputs")

col1, col2 = st.columns(2)

with col1:

    current_month_generation = st.number_input(
        "Solar Generation (kWh)",
        min_value=0.0,
        value=0.0
    )

    a_zone = st.number_input(
        "A Zone Units",
        min_value=0.0,
        value=0.0
    )

    b_zone = st.number_input(
        "B Zone Units",
        min_value=0.0,
        value=0.0
    )

    debit_bill_adjustment = st.number_input(
        "Debit Bill Adjustment (₹)",
        value=0.0
    )

with col2:

    c_zone = st.number_input(
        "C Zone Units",
        min_value=0.0,
        value=0.0
    )

    d_zone = st.number_input(
        "D Zone Units",
        min_value=0.0,
        value=0.0
    )

    grid_support_charges = st.number_input(
        "Grid Support Charges (₹)",
        value=0.0
    )

# ---------------------------------------------------
# FILE UPLOAD
# ---------------------------------------------------

uploaded_file = st.file_uploader(
    "Upload Electricity Bill PDF",
    type=["pdf"]
)

# ---------------------------------------------------
# CLEAN FUNCTION
# ---------------------------------------------------

def clean_number(value):

    if value:

        return (
            str(value)
            .replace(",", "")
            .replace("%", "")
            .replace("₹", "")
            .strip()
        )

    return "0"

# ---------------------------------------------------
# SAFE FLOAT
# ---------------------------------------------------

def safe_float(value):

    try:
        return float(clean_number(value))
    except:
        return 0

# ---------------------------------------------------
# MAIN PROCESS
# ---------------------------------------------------

if uploaded_file:

    text = ""

    try:

        # ---------------------------------------------------
        # READ PDF
        # ---------------------------------------------------

        with pdfplumber.open(uploaded_file) as pdf:

            for page in pdf.pages:

                extracted = page.extract_text()

                if extracted:
                    text += extracted + "\n"

        st.success("✅ PDF Processed Successfully")

        # ---------------------------------------------------
        # DEBUG TEXT
        # ---------------------------------------------------

        # st.text(text)

        # ---------------------------------------------------
        # MONTH
        # ---------------------------------------------------

        month_match = re.search(
            r'(Jan|Feb|Mar|Apr|May|Jun|Jul|Aug|Sep|Oct|Nov|Dec)[-\s]?\d{2}',
            text,
            re.IGNORECASE
        )

        bill_month = (
            month_match.group(0)
            if month_match else "May-26"
        )

        # ---------------------------------------------------
        # CONTRACT DEMAND
        # ---------------------------------------------------

        contract_match = re.search(
            r'Total\s*Contract\s*Demand\s*\(KVA\)\s*([\d,\.]+)',
            text,
            re.IGNORECASE
        )

        contract_demand = (
            contract_match.group(1)
            if contract_match else "0"
        )

        # ---------------------------------------------------
        # MAXIMUM DEMAND
        # ---------------------------------------------------

        md_match = re.search(
            r'Highest\s*Recorded\s*MSEDCL\s*Demand\s*([\d,\.]+)',
            text,
            re.IGNORECASE
        )

        if not md_match:

            md_match = re.search(
                r'MSEDCL\s*Demand\s*([\d,\.]+)',
                text,
                re.IGNORECASE
            )

        maximum_demand = (
            md_match.group(1)
            if md_match else "0"
        )

        # ---------------------------------------------------
        # TRANSMISSION CHARGES
        # ---------------------------------------------------

        tc_match = re.search(
            r'Transmission\s*Charges.*?([\d,]+\.\d+)',
            text,
            re.IGNORECASE
        )

        transmission_charges = (
            tc_match.group(1)
            if tc_match else "0"
        )

        # ---------------------------------------------------
        # BILLED DEMAND
        # ---------------------------------------------------

        billed_match = re.search(
            r'Billed\s*Demand\s*([\d,\.]+)',
            text,
            re.IGNORECASE
        )

        billed_demand = (
            billed_match.group(1)
            if billed_match else "0"
        )

        # ---------------------------------------------------
        # REFERENCE UNITS
        # ---------------------------------------------------

        ref_match = re.search(
            r'Ref\s*consumption\s*:?\s*([\d,\.]+)',
            text,
            re.IGNORECASE
        )

        reference_units = (
            ref_match.group(1)
            if ref_match else "0"
        )

        # ---------------------------------------------------
        # POWER FACTOR EXTRACTION
        # ---------------------------------------------------

        power_factor = "1"

        pf_patterns = [

            r'P\.?\s*F\.?\s*[:=\-]?\s*(0?\.\d+)',

            r'P\.?\s*F\.?\s*[:=\-]?\s*(1\.0+)',

            r'Power\s*Factor\s*[:=\-]?\s*(0?\.\d+)',

            r'Average\s*Power\s*Factor\s*[:=\-]?\s*(0?\.\d+)',

            r'Avg\.?\s*P\.?\s*F\.?\s*[:=\-]?\s*(0?\.\d+)'
        ]

        for pattern in pf_patterns:

            pf_match = re.search(
                pattern,
                text,
                re.IGNORECASE
            )

            if pf_match:

                power_factor = pf_match.group(1)
                break

        # ---------------------------------------------------
        # EXTRA PF SEARCH
        # ---------------------------------------------------

        if power_factor == "1":

            lines = text.split("\n")

            for line in lines:

                if "P.F" in line.upper() or "POWER FACTOR" in line.upper():

                    numbers = re.findall(
                        r'0\.\d+|1\.0+',
                        line
                    )

                    if numbers:

                        power_factor = numbers[0]
                        break

        # ---------------------------------------------------
        # ELECTRICITY DUTY
        # ---------------------------------------------------

        ed_match = re.search(
            r'Electricity\s*Duty\s*[:\-]?\s*([\d\.]+%)',
            text,
            re.IGNORECASE
        )

        electricity_duty = (
            ed_match.group(1)
            if ed_match else "7.50%"
        )

        # ---------------------------------------------------
        # STATIC VALUES
        # ---------------------------------------------------

        energy_rate = 8.44
        demand_charge_rate = 650
        wheeling_charge_rate = 0.81
        fac_rate = 0.50
        tax_rate = 0.29

        # ---------------------------------------------------
        # DISPLAY DATA
        # ---------------------------------------------------

        st.markdown("## 📋 Extracted Bill Data")

        col3, col4 = st.columns(2)

        with col3:

            st.write("Month:", bill_month)
            st.write("Contract Demand:", contract_demand)
            st.write("Maximum Demand:", maximum_demand)
            st.write("Transmission Charges:", transmission_charges)

        with col4:

            st.write("P.F.:", power_factor)
            st.write("Electricity Duty:", electricity_duty)
            st.write("Billed Demand:", billed_demand)
            st.write("Reference Units:", reference_units)

        st.markdown("---")

        # ---------------------------------------------------
        # GENERATE EXCEL REPORT
        # ---------------------------------------------------

        if st.button("Generate Excel Report"):

            try:

                # ---------------------------------------------------
                # TEMPLATE PATH
                # ---------------------------------------------------

                template_path = os.path.join(
                    "templates",
                    "bill_template.xlsx"
                )

                # ---------------------------------------------------
                # LOAD WORKBOOK
                # ---------------------------------------------------

                wb = load_workbook(template_path)

                # ---------------------------------------------------
                # SELECT SHEETS
                # ---------------------------------------------------

                input_sheet = wb[wb.sheetnames[0]]

                if len(wb.sheetnames) > 1:

                    output_sheet = wb[wb.sheetnames[1]]

                else:

                    output_sheet = input_sheet

                # ---------------------------------------------------
                # INPUT VALUES
                # ---------------------------------------------------

                input_sheet["C2"] = solar_capacity
                input_sheet["C3"] = plant_load

                input_sheet["C13"] = bill_month

                input_sheet["C14"] = safe_float(contract_demand)

                input_sheet["C15"] = energy_rate
                input_sheet["C16"] = demand_charge_rate
                input_sheet["C17"] = wheeling_charge_rate
                input_sheet["C18"] = fac_rate
                input_sheet["C19"] = tax_rate

                # ---------------------------------------------------
                # IMPORTANT PF CELL
                # ---------------------------------------------------

                input_sheet["C20"] = safe_float(power_factor)

                # ---------------------------------------------------
                # MAXIMUM DEMAND
                # ---------------------------------------------------

                input_sheet["C21"] = safe_float(maximum_demand)

                # ---------------------------------------------------
                # ELECTRICITY DUTY
                # ---------------------------------------------------

                input_sheet["C22"] = electricity_duty

                # ---------------------------------------------------
                # TRANSMISSION CHARGES
                # ---------------------------------------------------

                input_sheet["C9"] = safe_float(
                    transmission_charges
                )

                # ---------------------------------------------------
                # SOLAR GENERATION
                # ---------------------------------------------------

                input_sheet["H25"] = current_month_generation

                # ---------------------------------------------------
                # TOD ZONES
                # ---------------------------------------------------

                input_sheet["K26"] = a_zone
                input_sheet["L26"] = b_zone
                input_sheet["M26"] = c_zone
                input_sheet["N26"] = d_zone

                # ---------------------------------------------------
                # OTHER BILL VALUES
                # ---------------------------------------------------

                input_sheet["C30"] = safe_float(
                    billed_demand
                )

                input_sheet["C40"] = safe_float(
                    reference_units
                )

                # ---------------------------------------------------
                # OUTPUT VALUES
                # ---------------------------------------------------

                output_sheet["C22"] = (
                    debit_bill_adjustment
                )

                output_sheet["D22"] = (
                    debit_bill_adjustment
                    +
                    grid_support_charges
                )

                # ---------------------------------------------------
                # RECALCULATE
                # ---------------------------------------------------

                wb.calculation.fullCalcOnLoad = True
                wb.calculation.forceFullCalc = True

                # ---------------------------------------------------
                # SAVE OUTPUT
                # ---------------------------------------------------

                output = BytesIO()

                wb.save(output)

                output.seek(0)

                st.success(
                    "✅ Report Generated Successfully"
                )

                # ---------------------------------------------------
                # DOWNLOAD
                # ---------------------------------------------------

                st.download_button(
                    label="⬇ Download Excel Report",
                    data=output,
                    file_name="Before_After_Solar_Report.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )

            except Exception as excel_error:

                st.error(
                    f"Excel Generation Error: {excel_error}"
                )

    except Exception as e:

        st.error(
            f"PDF Processing Error: {e}"
        )
